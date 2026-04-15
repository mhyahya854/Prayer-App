#!/usr/bin/env python3
"""Build a unified hadith bundle from raw books and HadeethEnc sheets."""

from __future__ import annotations

import argparse
import hashlib
import json
import re
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


RAW_HADITH_BOOKS = [
    "bukhari",
    "muslim",
    "abudawud",
    "tirmidhi",
    "nasai",
    "ibnmajah",
]


@dataclass
class HadeethEncRow:
    id: str
    hadith_text: str
    title: str
    explanation: str
    explanation_ar: str
    benefits: str
    benefits_ar: str
    word_meanings: str
    grade: str
    grade_ar: str
    takhrij: str
    takhrij_ar: str
    link: str


def normalize_text(value: str) -> str:
    compact = re.sub(r"\s+", " ", value or "").strip()
    return compact


def text_fingerprint(value: str) -> str:
    normalized = normalize_text(value)
    if not normalized:
        return ""
    return hashlib.sha1(normalized.encode("utf-8")).hexdigest()


def to_list(value: str) -> list[str]:
    normalized = normalize_text(value)
    if not normalized:
        return []
    if "\n" in value:
        parts = [normalize_text(part) for part in value.splitlines()]
    else:
        parts = [normalize_text(part) for part in re.split(r"[;•]", value)]
    return [part for part in parts if part]


def load_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def load_hadeethenc_english(path: Path) -> dict[str, dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    header_values = list(next(sheet.iter_rows(min_row=2, max_row=2, values_only=True)))
    index_by_name = {str(value): idx for idx, value in enumerate(header_values) if value}
    rows: dict[str, dict[str, Any]] = {}

    for raw_row in sheet.iter_rows(min_row=3, values_only=True):
        row_id = str(raw_row[index_by_name["id"]]).strip() if raw_row[index_by_name["id"]] else ""
        if not row_id:
            continue
        rows[row_id] = {
            "id": row_id,
            "title": str(raw_row[index_by_name["title"]] or ""),
            "hadith_text": str(raw_row[index_by_name["hadith_text"]] or ""),
            "explanation": str(raw_row[index_by_name["explanation"]] or ""),
            "benefits": str(raw_row[index_by_name["benefits"]] or ""),
            "grade": str(raw_row[index_by_name["grade"]] or ""),
            "takhrij": str(raw_row[index_by_name["takhrij"]] or ""),
            "link": str(raw_row[index_by_name["link"]] or ""),
            "title_ar": str(raw_row[index_by_name["title_ar"]] or ""),
            "hadith_text_ar": str(raw_row[index_by_name["hadith_text_ar"]] or ""),
            "explanation_ar": str(raw_row[index_by_name["explanation_ar"]] or ""),
            "benefits_ar": str(raw_row[index_by_name["benefits_ar"]] or ""),
            "grade_ar": str(raw_row[index_by_name["grade_ar"]] or ""),
            "takhrij_ar": str(raw_row[index_by_name["takhrij_ar"]] or ""),
        }
    return rows


def load_hadeethenc_arabic(path: Path) -> dict[str, dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    header_values = list(next(sheet.iter_rows(min_row=2, max_row=2, values_only=True)))
    index_by_name = {str(value): idx for idx, value in enumerate(header_values) if value}
    rows: dict[str, dict[str, Any]] = {}

    for raw_row in sheet.iter_rows(min_row=3, values_only=True):
        row_id = str(raw_row[index_by_name["id"]]).strip() if raw_row[index_by_name["id"]] else ""
        if not row_id:
            continue
        rows[row_id] = {
            "id": row_id,
            "title": str(raw_row[index_by_name["title"]] or ""),
            "hadith_text": str(raw_row[index_by_name["hadith_text"]] or ""),
            "explanation": str(raw_row[index_by_name["explanation"]] or ""),
            "word_meanings": str(raw_row[index_by_name["word_meanings"]] or ""),
            "benefits": str(raw_row[index_by_name["benefits"]] or ""),
            "grade": str(raw_row[index_by_name["grade"]] or ""),
            "takhrij": str(raw_row[index_by_name["takhrij"]] or ""),
            "link": str(raw_row[index_by_name["link"]] or ""),
        }
    return rows


def merge_hadeethenc_rows(en_rows: dict[str, dict[str, Any]], ar_rows: dict[str, dict[str, Any]]) -> list[HadeethEncRow]:
    row_ids = sorted(set(en_rows.keys()) | set(ar_rows.keys()), key=lambda value: int(value))
    merged_rows: list[HadeethEncRow] = []

    for row_id in row_ids:
        en_row = en_rows.get(row_id, {})
        ar_row = ar_rows.get(row_id, {})
        merged_rows.append(
            HadeethEncRow(
                id=row_id,
                hadith_text=en_row.get("hadith_text", "") or ar_row.get("hadith_text", ""),
                title=en_row.get("title", "") or ar_row.get("title", ""),
                explanation=en_row.get("explanation", ""),
                explanation_ar=en_row.get("explanation_ar", "") or ar_row.get("explanation", ""),
                benefits=en_row.get("benefits", ""),
                benefits_ar=en_row.get("benefits_ar", "") or ar_row.get("benefits", ""),
                word_meanings=ar_row.get("word_meanings", ""),
                grade=en_row.get("grade", "") or ar_row.get("grade", ""),
                grade_ar=en_row.get("grade_ar", "") or ar_row.get("grade", ""),
                takhrij=en_row.get("takhrij", "") or ar_row.get("takhrij", ""),
                takhrij_ar=en_row.get("takhrij_ar", "") or ar_row.get("takhrij", ""),
                link=en_row.get("link", "") or ar_row.get("link", ""),
            )
        )
    return merged_rows


def build_hadith_bundle(project_root: Path) -> dict[str, Any]:
    raw_hadith_dir = project_root / "content-src" / "raw" / "hadith"
    raw_hadeethenc_dir = project_root / "content-src" / "raw" / "hadeethenc"
    generated_dir = project_root / "content-src" / "generated"
    generated_dir.mkdir(parents=True, exist_ok=True)

    en_rows = load_hadeethenc_english(raw_hadeethenc_dir / "hadeethenc_en.xlsx")
    ar_rows = load_hadeethenc_arabic(raw_hadeethenc_dir / "hadeethenc_ar.xlsx")
    merged_rows = merge_hadeethenc_rows(en_rows, ar_rows)

    by_id = {row.id: row for row in merged_rows}
    by_fingerprint = {text_fingerprint(row.hadith_text): row for row in merged_rows if row.hadith_text}

    books: list[dict[str, Any]] = []
    chapters: list[dict[str, Any]] = []
    entries: list[dict[str, Any]] = []
    hadith_sources: dict[str, int] = {}
    matched_count = 0

    for book_slug in RAW_HADITH_BOOKS:
        book_path = raw_hadith_dir / f"{book_slug}.json"
        raw_book = load_json(book_path)
        metadata = raw_book.get("metadata", {})
        book_id = int(raw_book["id"])

        books.append(
            {
                "id": book_id,
                "slug": book_slug,
                "titleArabic": metadata.get("arabic", {}).get("title", ""),
                "titleEnglish": metadata.get("english", {}).get("title", ""),
                "authorArabic": metadata.get("arabic", {}).get("author", ""),
                "authorEnglish": metadata.get("english", {}).get("author", ""),
                "hadithCount": int(metadata.get("length") or len(raw_book.get("hadiths", []))),
                "isShipNow": book_slug in {"bukhari", "muslim"},
            }
        )

        chapter_by_id: dict[int, dict[str, Any]] = {}
        for chapter in raw_book.get("chapters", []):
            if chapter.get("id") is None:
                continue
            chapter_record = {
                "id": int(chapter["id"]),
                "bookId": book_id,
                "bookSlug": book_slug,
                "titleArabic": chapter.get("arabic", ""),
                "titleEnglish": chapter.get("english", ""),
            }
            chapter_by_id[chapter_record["id"]] = chapter_record
            chapters.append(chapter_record)

        for hadith in raw_book.get("hadiths", []):
            if hadith.get("idInBook") is None or hadith.get("chapterId") is None:
                continue
            english_payload = hadith.get("english") or {}
            narrator = ""
            english_text = ""
            if isinstance(english_payload, dict):
                narrator = str(english_payload.get("narrator") or "")
                english_text = str(english_payload.get("text") or "")
            elif isinstance(english_payload, str):
                english_text = english_payload

            chapter_id = int(hadith["chapterId"])
            chapter = chapter_by_id.get(chapter_id)
            row = by_id.get(str(hadith["idInBook"])) or by_fingerprint.get(text_fingerprint(english_text))
            if row:
                matched_count += 1

            entry_id = f"{book_slug}:{hadith['idInBook']}"
            hadith_sources[book_slug] = hadith_sources.get(book_slug, 0) + 1

            entries.append(
                {
                    "id": entry_id,
                    "bookId": book_id,
                    "bookSlug": book_slug,
                    "bookHadithNumber": int(hadith["idInBook"]),
                    "globalHadithId": int(hadith["id"]),
                    "chapterId": chapter_id,
                    "chapterTitleArabic": chapter["titleArabic"] if chapter else "",
                    "chapterTitleEnglish": chapter["titleEnglish"] if chapter else "",
                    "narratorEnglish": narrator,
                    "textArabic": str(hadith.get("arabic") or ""),
                    "textEnglish": english_text,
                    "grade": row.grade if row else "",
                    "gradeArabic": row.grade_ar if row else "",
                    "explanation": row.explanation if row else "",
                    "explanationArabic": row.explanation_ar if row else "",
                    "benefits": to_list(row.benefits) if row else [],
                    "benefitsArabic": to_list(row.benefits_ar) if row else [],
                    "wordMeaningsArabic": to_list(row.word_meanings) if row else [],
                    "takhrij": row.takhrij if row else "",
                    "takhrijArabic": row.takhrij_ar if row else "",
                    "sourceLink": row.link if row else "",
                    "hadeethEncId": row.id if row else None,
                    "isGradeVerified": bool(row and normalize_text(row.grade)),
                }
            )

    bundle = {
        "source": {
            "collection": "Hadith Corpus + HadeethEnc",
            "license": "See original source licenses",
            "name": "Prayer App Unified Hadith",
            "version": datetime.now(tz=timezone.utc).strftime("%Y.%m.%d"),
            "generatedAt": datetime.now(tz=timezone.utc).isoformat(),
            "books": RAW_HADITH_BOOKS,
        },
        "stats": {
            "bookCount": len(books),
            "chapterCount": len(chapters),
            "entryCount": len(entries),
            "matchedHadeethEncCount": matched_count,
            "unmatchedCount": max(len(entries) - matched_count, 0),
            "entriesByBook": hadith_sources,
        },
        "books": books,
        "chapters": chapters,
        "entries": entries,
    }

    output_path = generated_dir / "hadith.bundle.json"
    output_path.write_text(json.dumps(bundle, ensure_ascii=False, indent=2), encoding="utf-8")
    return bundle


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build hadith bundle JSON from raw sources.")
    parser.add_argument(
        "--project-root",
        default=Path(__file__).resolve().parents[1],
        type=Path,
        help="Project root path.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    bundle = build_hadith_bundle(args.project_root.resolve())
    print(
        f"Built hadith bundle with {bundle['stats']['entryCount']} entries "
        f"({bundle['stats']['matchedHadeethEncCount']} matched to HadeethEnc)."
    )


if __name__ == "__main__":
    main()
